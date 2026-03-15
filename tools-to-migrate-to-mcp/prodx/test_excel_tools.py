"""
Comprehensive tests for Excel Tools

Tests all 4 Excel tools:
- ReadExcelStructuredTool
- ModifyExcelTool
- CreateExcelChartTool
- AnalyzeExcelSecurityTool
"""

import pytest
import base64
import io
import json
import tempfile
import os

# Conditional imports for optional dependencies
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    pytest.skip("openpyxl not available", allow_module_level=True)

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pytest.skip("pandas not available", allow_module_level=True)

# Import tools - handle both relative and absolute imports
try:
    from .excel_tools import (
        ReadExcelStructuredTool,
        ModifyExcelTool,
        CreateExcelChartTool,
        AnalyzeExcelSecurityTool,
        _decode_file_input,
        _encode_file_output
    )
except ImportError:
    # Fallback for direct execution
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))))
    from shared.modules.tools.prodx.excel_tools import (
        ReadExcelStructuredTool,
        ModifyExcelTool,
        CreateExcelChartTool,
        AnalyzeExcelSecurityTool,
        _decode_file_input,
        _encode_file_output
    )


class TestHelperFunctions:
    """Test helper functions for file encoding/decoding."""
    
    def test_decode_file_input_base64(self):
        """Test decoding base64 file input."""
        test_data = b"test excel content"
        encoded = base64.b64encode(test_data).decode('utf-8')
        
        file_bytes, is_base64 = _decode_file_input(encoded)
        assert file_bytes == test_data
        assert is_base64 is True
    
    def test_decode_file_input_file_path(self):
        """Test decoding file path input."""
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(b"test content")
            tmp_path = tmp.name
        
        try:
            file_bytes, is_base64 = _decode_file_input(tmp_path)
            assert file_bytes == b"test content"
            assert is_base64 is False
        finally:
            os.unlink(tmp_path)
    
    def test_decode_file_input_invalid(self):
        """Test decoding invalid input."""
        with pytest.raises(ValueError):
            _decode_file_input("/nonexistent/file.xlsx")
    
    def test_encode_file_output(self):
        """Test encoding file bytes to base64."""
        test_data = b"test content"
        encoded = _encode_file_output(test_data)
        decoded = base64.b64decode(encoded)
        assert decoded == test_data


class TestReadExcelStructuredTool:
    """Test ReadExcelStructuredTool."""
    
    def create_test_excel(self, sheet_name="Sheet1", data=None):
        """Create a test Excel file in memory."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if data is None:
            data = [
                ["Name", "Age", "City"],
                ["Alice", 30, "New York"],
                ["Bob", 25, "London"],
                ["Charlie", 35, "Tokyo"]
            ]
        
        for row in data:
            ws.append(row)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return base64.b64encode(output.read()).decode('utf-8')
    
    def test_read_excel_basic(self):
        """Test basic Excel reading."""
        tool = ReadExcelStructuredTool()
        excel_data = self.create_test_excel()
        
        result = tool._run(excel_data)
        result_dict = json.loads(result)
        
        assert result_dict["status"] == "success"
        assert "data" in result_dict
        assert len(result_dict["data"]) > 0
        assert "Sheet1" in result_dict["sheets"]
    
    def test_read_excel_with_sheet_name(self):
        """Test reading specific sheet."""
        tool = ReadExcelStructuredTool()
        excel_data = self.create_test_excel(sheet_name="Data")
        
        result = tool._run(excel_data, sheet_name="Data")
        result_dict = json.loads(result)
        
        assert result_dict["status"] == "success"
        assert "Data" in result_dict["sheets"]
    
    def test_read_excel_with_formulas(self):
        """Test reading Excel with formulas."""
        tool = ReadExcelStructuredTool()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 10
        ws['B1'] = 20
        ws['C1'] = '=A1+B1'
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = base64.b64encode(output.read()).decode('utf-8')
        
        result = tool._run(excel_data, include_formulas=True)
        result_dict = json.loads(result)
        
        assert result_dict["status"] == "success"
        # Check if formulas are included
        data = result_dict["data"]
        formula_found = any("=A1+B1" in str(cell.get("formula", "")) for cell in data)
        assert formula_found
    
    def test_read_excel_invalid_base64(self):
        """Test reading with invalid base64."""
        tool = ReadExcelStructuredTool()
        result = tool._run("invalid_base64_data")
        assert "Error" in result
    
    def test_read_excel_empty_file(self):
        """Test reading empty Excel file."""
        tool = ReadExcelStructuredTool()
        wb = Workbook()
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = base64.b64encode(output.read()).decode('utf-8')
        
        result = tool._run(excel_data)
        result_dict = json.loads(result)
        assert result_dict["status"] == "success"


class TestModifyExcelTool:
    """Test ModifyExcelTool."""
    
    def create_test_excel(self):
        """Create a test Excel file."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Name"
        ws['B1'] = "Value"
        ws['A2'] = "Item1"
        ws['B2'] = 100
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return base64.b64encode(output.read()).decode('utf-8')
    
    def test_modify_excel_update_cell(self):
        """Test updating a cell value."""
        tool = ModifyExcelTool()
        excel_data = self.create_test_excel()
        
        operations = [
            {
                "type": "update_cell",
                "sheet": "Sheet1",
                "cell": "B2",
                "value": 200
            }
        ]
        
        result = tool._run(excel_data, operations)
        result_dict = json.loads(result)
        
        assert result_dict["status"] == "success"
        assert "file_data" in result_dict
    
    def test_modify_excel_add_row(self):
        """Test adding a row."""
        tool = ModifyExcelTool()
        excel_data = self.create_test_excel()
        
        operations = [
            {
                "type": "add_row",
                "sheet": "Sheet1",
                "row_data": ["Item2", 300]
            }
        ]
        
        result = tool._run(excel_data, operations)
        result_dict = json.loads(result)
        assert result_dict["status"] == "success"
    
    def test_modify_excel_add_formula(self):
        """Test adding a formula."""
        tool = ModifyExcelTool()
        excel_data = self.create_test_excel()
        
        operations = [
            {
                "type": "update_cell",
                "sheet": "Sheet1",
                "cell": "B3",
                "value": "=SUM(B2:B2)"
            }
        ]
        
        result = tool._run(excel_data, operations)
        result_dict = json.loads(result)
        assert result_dict["status"] == "success"
    
    def test_modify_excel_multiple_operations(self):
        """Test multiple operations."""
        tool = ModifyExcelTool()
        excel_data = self.create_test_excel()
        
        operations = [
            {
                "type": "update_cell",
                "sheet": "Sheet1",
                "cell": "A3",
                "value": "Item2"
            },
            {
                "type": "update_cell",
                "sheet": "Sheet1",
                "cell": "B3",
                "value": 200
            }
        ]
        
        result = tool._run(excel_data, operations)
        result_dict = json.loads(result)
        assert result_dict["status"] == "success"
    
    def test_modify_excel_invalid_operation(self):
        """Test invalid operation type."""
        tool = ModifyExcelTool()
        excel_data = self.create_test_excel()
        
        operations = [
            {
                "type": "invalid_operation",
                "sheet": "Sheet1"
            }
        ]
        
        result = tool._run(excel_data, operations)
        assert "Error" in result
    
    def test_modify_excel_invalid_cell(self):
        """Test invalid cell reference."""
        tool = ModifyExcelTool()
        excel_data = self.create_test_excel()
        
        operations = [
            {
                "type": "update_cell",
                "sheet": "Sheet1",
                "cell": "ZZ999",
                "value": 100
            }
        ]
        
        result = tool._run(excel_data, operations)
        # Should handle gracefully
        assert isinstance(result, str)


class TestCreateExcelChartTool:
    """Test CreateExcelChartTool."""
    
    def create_test_excel(self):
        """Create a test Excel file with data."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Category", "Value"])
        ws.append(["A", 10])
        ws.append(["B", 20])
        ws.append(["C", 30])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return base64.b64encode(output.read()).decode('utf-8')
    
    def test_create_bar_chart(self):
        """Test creating a bar chart."""
        tool = CreateExcelChartTool()
        excel_data = self.create_test_excel()
        
        chart_config = {
            "chart_type": "bar",
            "data_range": "A1:B4",
            "title": "Test Chart",
            "x_axis_title": "Category",
            "y_axis_title": "Value"
        }
        
        result = tool._run(excel_data, chart_config)
        result_dict = json.loads(result)
        assert result_dict["status"] == "success"
    
    def test_create_line_chart(self):
        """Test creating a line chart."""
        tool = CreateExcelChartTool()
        excel_data = self.create_test_excel()
        
        chart_config = {
            "chart_type": "line",
            "data_range": "A1:B4",
            "title": "Line Chart"
        }
        
        result = tool._run(excel_data, chart_config)
        result_dict = json.loads(result)
        assert result_dict["status"] == "success"
    
    def test_create_pie_chart(self):
        """Test creating a pie chart."""
        tool = CreateExcelChartTool()
        excel_data = self.create_test_excel()
        
        chart_config = {
            "chart_type": "pie",
            "data_range": "A1:B4",
            "title": "Pie Chart"
        }
        
        result = tool._run(excel_data, chart_config)
        result_dict = json.loads(result)
        assert result_dict["status"] == "success"
    
    def test_create_chart_invalid_range(self):
        """Test creating chart with invalid range."""
        tool = CreateExcelChartTool()
        excel_data = self.create_test_excel()
        
        chart_config = {
            "chart_type": "bar",
            "data_range": "INVALID",
            "title": "Test"
        }
        
        result = tool._run(excel_data, chart_config)
        assert isinstance(result, str)


class TestAnalyzeExcelSecurityTool:
    """Test AnalyzeExcelSecurityTool."""
    
    def create_test_excel(self, with_macro=False):
        """Create a test Excel file."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Test"
        ws['B1'] = "Data"
        
        if with_macro:
            # Note: openpyxl doesn't support VBA macros directly
            # This is a placeholder for testing
            pass
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return base64.b64encode(output.read()).decode('utf-8')
    
    def test_analyze_security_basic(self):
        """Test basic security analysis."""
        tool = AnalyzeExcelSecurityTool()
        excel_data = self.create_test_excel()
        
        result = tool._run(excel_data)
        result_dict = json.loads(result)
        
        assert result_dict["status"] == "success"
        assert "security_issues" in result_dict
        assert isinstance(result_dict["security_issues"], list)
    
    def test_analyze_security_with_external_links(self):
        """Test security analysis with external links."""
        tool = AnalyzeExcelSecurityTool()
        excel_data = self.create_test_excel()
        
        result = tool._run(excel_data)
        result_dict = json.loads(result)
        
        assert result_dict["status"] == "success"
        # Should detect or report on external links
    
    def test_analyze_security_invalid_file(self):
        """Test security analysis with invalid file."""
        tool = AnalyzeExcelSecurityTool()
        
        result = tool._run("invalid_base64")
        assert "Error" in result


class TestRealWorldScenarios:
    """Real-world scenario tests for Excel tools."""
    
    def test_read_large_realistic_excel(self):
        """Test reading a realistic large Excel file."""
        tool = ReadExcelStructuredTool()
        
        # Create realistic Excel with multiple sheets
        wb = Workbook()
        
        # Sheet 1: Sales data
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.append(["Date", "Product", "Quantity", "Price", "Total"])
        for i in range(100):
            quantity = (i % 20) + 1
            price = (i % 100) + 10
            total = quantity * price
            ws1.append([f"2024-01-{(i % 28) + 1:02d}", f"Product_{i % 10}", quantity, price, total])
        
        # Sheet 2: Inventory
        ws2 = wb.create_sheet("Inventory")
        ws2.append(["Product", "Stock", "Location"])
        for i in range(50):
            ws2.append([f"Product_{i}", (i % 100) + 50, f"Warehouse_{i % 5}"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = base64.b64encode(output.read()).decode('utf-8')
        
        # Test reading
        result = tool._run(excel_data)
        result_dict = json.loads(result)
        assert result_dict["status"] == "success"
        assert len(result_dict.get("sheets", [])) >= 2
    
    def test_modify_excel_complex_operations(self):
        """Test complex modification operations."""
        tool = ModifyExcelTool()
        excel_data = self.create_test_excel()
        
        # Multiple operations
        operations = [
            {
                "type": "update_cell",
                "sheet": "Sheet1",
                "cell": "A1",
                "value": "Updated Header"
            },
            {
                "type": "update_cell",
                "sheet": "Sheet1",
                "cell": "B2",
                "value": 999
            },
            {
                "type": "add_row",
                "sheet": "Sheet1",
                "row_data": ["New", "Row", "Data", 100]
            },
            {
                "type": "update_cell",
                "sheet": "Sheet1",
                "cell": "D7",
                "value": "=SUM(D2:D6)"
            }
        ]
        
        result = tool._run(excel_data, operations)
        result_dict = json.loads(result)
        assert result_dict["status"] == "success"
    
    def test_chart_creation_realistic_data(self):
        """Test chart creation with realistic business data."""
        tool = CreateExcelChartTool()
        
        # Create Excel with time series data
        wb = Workbook()
        ws = wb.active
        ws.append(["Month", "Revenue", "Expenses", "Profit"])
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        for i, month in enumerate(months):
            revenue = 50000 + (i * 5000)
            expenses = 30000 + (i * 3000)
            profit = revenue - expenses
            ws.append([month, revenue, expenses, profit])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = base64.b64encode(output.read()).decode('utf-8')
        
        # Test different chart types
        chart_types = ["bar", "line", "pie"]
        for chart_type in chart_types:
            chart_config = {
                "chart_type": chart_type,
                "data_range": "A1:D13",
                "title": f"{chart_type.title()} Chart - Financial Data"
            }
            result = tool._run(excel_data, chart_config)
            result_dict = json.loads(result)
            assert result_dict["status"] == "success"


class TestErrorHandling:
    """Test error handling across all tools."""
    
    def test_all_tools_handle_missing_library(self):
        """Test that tools handle missing libraries gracefully."""
        # This would require mocking imports, but the tools should return error messages
        # instead of crashing
        pass
    
    def test_all_tools_handle_invalid_base64(self):
        """Test that all tools handle invalid base64."""
        tools = [
            ReadExcelStructuredTool(),
            ModifyExcelTool(),
            CreateExcelChartTool(),
            AnalyzeExcelSecurityTool()
        ]
        
        for tool in tools:
            # Each tool should return an error message, not crash
            result = tool._run("invalid_base64_data")
            assert isinstance(result, str)
            assert len(result) > 0
    
    def test_all_tools_handle_empty_input(self):
        """Test that all tools handle empty input."""
        tools = [
            ReadExcelStructuredTool(),
            ModifyExcelTool(),
            CreateExcelChartTool(),
            AnalyzeExcelSecurityTool()
        ]
        
        for tool in tools:
            result = tool._run("")
            assert isinstance(result, str)
    
    def test_all_tools_handle_corrupted_excel(self):
        """Test that tools handle corrupted Excel files."""
        tools = [
            ReadExcelStructuredTool(),
            ModifyExcelTool(),
            CreateExcelChartTool(),
            AnalyzeExcelSecurityTool()
        ]
        
        # Create corrupted base64 (valid base64 but not valid Excel)
        corrupted_data = base64.b64encode(b"not an excel file").decode('utf-8')
        
        for tool in tools:
            result = tool._run(corrupted_data)
            assert isinstance(result, str)
            # Should return error, not crash
            assert "Error" in result or json.loads(result).get("status") != "success"
    
    def test_all_tools_handle_very_large_files(self):
        """Test that tools handle very large files gracefully."""
        # Create a large but valid Excel file
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Data"])
        for i in range(5000):  # 5000 rows
            ws.append([i, f"Data row {i}"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        large_excel_data = base64.b64encode(output.read()).decode('utf-8')
        
        # Test reading (might be slow but should work)
        read_tool = ReadExcelStructuredTool()
        result = read_tool._run(large_excel_data)
        result_dict = json.loads(result)
        # Should either succeed or return a reasonable error
        assert isinstance(result_dict, dict) or "Error" in result


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])

