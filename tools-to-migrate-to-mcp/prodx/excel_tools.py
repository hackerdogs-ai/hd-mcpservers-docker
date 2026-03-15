"""
Excel Operations Tools
----------------------
LangChain tools for Excel file manipulation, chart creation, and security analysis.
"""

from langchain_core.tools import BaseTool
from pydantic import BaseModel, Field
from typing import Type, Dict, List, Optional, Any, Union
import sys
import os
# Add project root to path for shared modules
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../../../'))
try:
    from shared.logger import setup_logger
except ImportError:
    # Fallback for testing environments
    import logging
    def setup_logger(name, log_file_path=None, **kwargs):
        logger = logging.getLogger(name)
        if not logger.handlers:
            handler = logging.StreamHandler()
            handler.setFormatter(logging.Formatter('%(name)s - %(levelname)s - %(message)s'))
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        return logger
import base64
import io
import json

logger = setup_logger(__name__, log_file_path="logs/excel_tools.log")

# Try to import required libraries
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
    from openpyxl.chart.reference import Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel tools will not work.")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas not available. Some Excel tools will not work.")


def _decode_file_input(file_input: str) -> tuple[bytes, bool]:
    """
    Decode file input which can be either a base64 string or file path.
    Returns (file_bytes, is_base64)
    """
    logger.debug(f"Attempting to decode file input (length: {len(file_input) if file_input else 0})")
    
    try:
        # Try to decode as base64 first
        logger.debug("Attempting base64 decoding")
        file_bytes = base64.b64decode(file_input)
        logger.debug(f"Successfully decoded base64 input (size: {len(file_bytes)} bytes)")
        return file_bytes, True
    except Exception as base64_error:
        logger.debug(f"Base64 decoding failed: {str(base64_error)}, treating as file path")
        # If that fails, treat as file path
        try:
            logger.debug(f"Attempting to read file from path: {file_input[:100] if len(file_input) > 100 else file_input}")
            with open(file_input, 'rb') as f:
                file_bytes = f.read()
                logger.debug(f"Successfully read file from path (size: {len(file_bytes)} bytes)")
                return file_bytes, False
        except FileNotFoundError as fnf_error:
            logger.error(f"File not found: {file_input}")
            raise ValueError(f"File not found: {file_input}")
        except PermissionError as perm_error:
            logger.error(f"Permission denied reading file: {file_input}")
            raise ValueError(f"Permission denied reading file: {file_input}")
        except Exception as file_error:
            logger.error(f"Error reading file from path: {str(file_error)}", exc_info=True)
            raise ValueError(f"Invalid file input: {file_error}")


def _encode_file_output(file_bytes: bytes) -> str:
    """Encode file bytes to base64 string."""
    try:
        logger.debug(f"Encoding {len(file_bytes)} bytes to base64")
        encoded = base64.b64encode(file_bytes).decode('utf-8')
        logger.debug(f"Successfully encoded to base64 (length: {len(encoded)} chars)")
        return encoded
    except Exception as e:
        logger.error(f"Error encoding file bytes to base64: {str(e)}", exc_info=True)
        raise ValueError(f"Failed to encode file to base64: {str(e)}")


# --- Excel Tools ---

class ReadExcelStructuredInput(BaseModel):
    """Input schema for ReadExcelStructuredTool."""
    file_path: str = Field(..., description="Path to Excel file or base64-encoded file data")
    sheet_name: Optional[str] = Field(None, description="Specific sheet name to read (default: first sheet)")
    include_formulas: bool = Field(False, description="Include formulas in output (default: False)")
    include_formatting: bool = Field(False, description="Include formatting information (default: False)")


class ReadExcelStructuredTool(BaseTool):
    """
    Read Excel files while preserving structure (cells, formulas, formatting).
    
    This tool reads Excel files and returns structured data including cell values with coordinates,
    formulas (if requested), formatting information (if requested), and sheet names. Unlike simple
    text extraction, this tool maintains the Excel file structure, making it ideal for analysis that
    requires understanding cell relationships, formulas, and formatting.
    
    **When to use this tool:**
    - Analyzing Excel file structure and organization
    - Extracting data with cell references (e.g., "Cell A1 contains...")
    - Understanding formulas and their dependencies
    - Processing Excel files for security analysis (checking for malicious formulas)
    - Extracting formatted data while preserving formatting metadata
    - Analyzing spreadsheet structure before modification
    - Understanding data relationships in complex spreadsheets
    
    **When NOT to use this tool:**
    - For simple data extraction where structure doesn't matter (use regular file reading)
    - When you only need the raw text content (use document extraction tools)
    - For very large Excel files with thousands of cells (may be slow)
    - When you need to modify the file (use ModifyExcelTool instead)
    
    **Input requirements:**
    - Must provide a valid Excel file (.xlsx or .xls format)
    - File can be provided as base64-encoded string or file path
    - Sheet name is optional (defaults to first/active sheet)
    - include_formulas and include_formatting are optional flags (default: False)
    
    **Output:**
    - Returns structured JSON with:
      - Sheet name and dimensions (max_row, max_column)
      - List of cells with coordinates, values, and optional formulas/formatting
      - All sheet names in the workbook
    - Returns error message if file cannot be read or sheet not found
    
    **Limitations:**
    - Large files with many cells may take time to process
    - Formula extraction requires formulas to be present (not calculated values)
    - Formatting extraction adds processing overhead
    - Only supports .xlsx and .xls formats (not .csv)
    
    **Example use cases:**
    1. "Read the Excel file and show me all formulas in column C"
    2. "Analyze the structure of this spreadsheet and identify all cell references"
    3. "Extract all formatted cells (bold, colored) from the Excel file"
    4. "Read the 'Sales' sheet and show me the cell structure"
    
    **Configuration:**
    Requires openpyxl library to be installed. The tool automatically handles both base64-encoded
    files (from Streamlit uploads) and file paths (for local files).
    """
    name: str = "read_excel_structured"
    description: str = (
        "Read Excel files while preserving structure (cells, formulas, formatting). "
        "Returns structured JSON with cell data, coordinates, formulas, and formatting information. "
        "Use this when you need to analyze Excel file structure, extract data with cell references, "
        "understand formulas and their dependencies, or process Excel files for security analysis. "
        "IMPORTANT: File can be base64-encoded string or file path. Sheet name is optional (defaults to active sheet). "
        "Best for: structure analysis, formula extraction, security auditing, formatted data extraction. "
        "NOT suitable for: simple text extraction, very large files, or when you need to modify the file."
    )
    args_schema: Type[BaseModel] = ReadExcelStructuredInput

    def _run(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        include_formulas: bool = False,
        include_formatting: bool = False
    ) -> str:
        """
        Read Excel file with structure preservation.
        
        This method performs the actual Excel file reading operation:
        1. Decodes file input (base64 or file path)
        2. Loads workbook using openpyxl
        3. Selects specified sheet or active sheet
        4. Iterates through all cells extracting values, coordinates
        5. Optionally includes formulas and formatting information
        6. Returns structured JSON with all extracted data
        
        Args:
            file_path: Path to Excel file or base64-encoded file data.
                      Example: "data.xlsx" or base64 string from file upload
            sheet_name: Optional sheet name to read. If not provided, uses active sheet.
                       Example: "Sales", "Sheet1", etc.
            include_formulas: If True, includes cell formulas in output (default: False)
            include_formatting: If True, includes formatting info (font, fill, alignment) (default: False)
        
        Returns:
            str: JSON string containing structured Excel data with cells, coordinates,
                 optional formulas and formatting. Returns error message string if reading fails.
        
        Raises:
            No exceptions are raised - all errors are caught and returned as error message strings.
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl library not available")
            return "Error: openpyxl library is not installed. Please install it with: pip install openpyxl"
        
        try:
            logger.info(f"Starting Excel file read operation")
            logger.debug(f"Parameters: sheet_name={sheet_name}, include_formulas={include_formulas}, include_formatting={include_formatting}")
            logger.debug(f"File input length: {len(file_path) if file_path else 0}")
            
            # Decode file input
            try:
                logger.debug("Decoding file input")
                file_bytes, is_base64 = _decode_file_input(file_path)
                logger.info(f"File decoded successfully (size: {len(file_bytes)} bytes, is_base64: {is_base64})")
            except ValueError as decode_error:
                logger.error(f"File input decoding failed: {str(decode_error)}", exc_info=True)
                return f"Error: Invalid file input - {str(decode_error)}"
            except Exception as decode_error:
                logger.error(f"Unexpected error during file decoding: {str(decode_error)}", exc_info=True)
                return f"Error: Failed to decode file input - {str(decode_error)}"
            
            # Create BytesIO object
            try:
                logger.debug("Creating BytesIO object from file bytes")
                file_obj = io.BytesIO(file_bytes)
                logger.debug("BytesIO object created successfully")
            except Exception as io_error:
                logger.error(f"Error creating BytesIO object: {str(io_error)}", exc_info=True)
                return f"Error: Failed to create file object - {str(io_error)}"
            
            # Load workbook
            try:
                logger.debug(f"Loading workbook (data_only={not include_formulas})")
                wb = load_workbook(file_obj, data_only=not include_formulas)
                logger.info(f"Workbook loaded successfully (sheets: {len(wb.sheetnames)})")
                logger.debug(f"Available sheets: {', '.join(wb.sheetnames)}")
            except Exception as load_error:
                logger.error(f"Error loading workbook: {str(load_error)}", exc_info=True)
                return f"Error: Failed to load Excel workbook - {str(load_error)}"
            
            # Select sheet
            try:
                if sheet_name:
                    logger.debug(f"Selecting sheet: {sheet_name}")
                    if sheet_name not in wb.sheetnames:
                        logger.warning(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
                        return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
                    ws = wb[sheet_name]
                    logger.debug(f"Sheet '{sheet_name}' selected successfully")
                else:
                    logger.debug("No sheet name specified, using active sheet")
                    ws = wb.active
                    logger.debug(f"Active sheet selected: {ws.title}")
            except Exception as sheet_error:
                logger.error(f"Error selecting sheet: {str(sheet_error)}", exc_info=True)
                return f"Error: Failed to select sheet - {str(sheet_error)}"
            
            # Extract structured data
            try:
                logger.debug("Initializing result structure")
                result = {
                    "sheet_name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "cells": []
                }
                logger.debug(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
            except Exception as init_error:
                logger.error(f"Error initializing result structure: {str(init_error)}", exc_info=True)
                return f"Error: Failed to initialize result structure - {str(init_error)}"
            
            # Read cells
            try:
                logger.debug("Starting cell iteration")
                cells_processed = 0
                cells_with_data = 0
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1):
                    for cell in row:
                        cells_processed += 1
                        try:
                            if cell.value is not None:
                                cells_with_data += 1
                                cell_data = {
                                    "coordinate": cell.coordinate,
                                    "row": cell.row,
                                    "column": cell.column,
                                    "value": str(cell.value)
                                }
                                
                                if include_formulas and cell.data_type == 'f':
                                    try:
                                        cell_data["formula"] = cell.value
                                        logger.debug(f"Added formula for cell {cell.coordinate}")
                                    except Exception as formula_error:
                                        logger.warning(f"Error extracting formula from cell {cell.coordinate}: {str(formula_error)}")
                                
                                if include_formatting:
                                    try:
                                        cell_data["formatting"] = {
                                            "font": {
                                                "name": cell.font.name if cell.font else None,
                                                "size": cell.font.size if cell.font else None,
                                                "bold": cell.font.bold if cell.font else False,
                                                "italic": cell.font.italic if cell.font else False,
                                                "color": str(cell.font.color.rgb) if cell.font and cell.font.color else None
                                            },
                                            "fill": {
                                                "fgColor": str(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor else None
                                            },
                                            "alignment": {
                                                "horizontal": cell.alignment.horizontal if cell.alignment else None,
                                                "vertical": cell.alignment.vertical if cell.alignment else None
                                            }
                                        }
                                        logger.debug(f"Added formatting for cell {cell.coordinate}")
                                    except Exception as format_error:
                                        logger.warning(f"Error extracting formatting from cell {cell.coordinate}: {str(format_error)}")
                                
                                result["cells"].append(cell_data)
                        except Exception as cell_error:
                            logger.warning(f"Error processing cell {cell.coordinate if 'cell' in locals() else 'unknown'}: {str(cell_error)}")
                            continue  # Skip this cell and continue
                
                logger.debug(f"Cell iteration complete: {cells_processed} cells processed, {cells_with_data} cells with data")
            except Exception as cell_iter_error:
                logger.error(f"Error during cell iteration: {str(cell_iter_error)}", exc_info=True)
                return f"Error: Failed to read cells - {str(cell_iter_error)}"
            
            # Add sheet names
            try:
                result["all_sheets"] = wb.sheetnames
                logger.debug(f"Added sheet names: {len(wb.sheetnames)} sheets")
            except Exception as sheet_names_error:
                logger.warning(f"Error adding sheet names: {str(sheet_names_error)}")
                result["all_sheets"] = []
            
            # Serialize to JSON
            try:
                logger.debug("Serializing result to JSON")
                json_result = json.dumps(result, indent=2)
                logger.info(f"Successfully read Excel file: {len(result['cells'])} cells extracted from sheet '{result['sheet_name']}'")
                logger.debug(f"JSON result length: {len(json_result)} characters")
                return json_result
            except Exception as json_error:
                logger.error(f"Error serializing result to JSON: {str(json_error)}", exc_info=True)
                return f"Error: Failed to serialize result - {str(json_error)}"
            
        except ValueError as val_error:
            logger.error(f"Value error in Excel read operation: {str(val_error)}", exc_info=True)
            return f"Error: Invalid input - {str(val_error)}"
        except json.JSONDecodeError as json_error:
            logger.error(f"JSON error in Excel read operation: {str(json_error)}", exc_info=True)
            return f"Error: JSON serialization failed - {str(json_error)}"
        except MemoryError as mem_error:
            logger.error(f"Memory error reading Excel file (file may be too large): {str(mem_error)}", exc_info=True)
            return f"Error: File too large to process - {str(mem_error)}"
        except Exception as e:
            logger.error(f"Unexpected error reading Excel file: {str(e)}", exc_info=True)
            return f"Error reading Excel file: {str(e)}"
    
    async def _arun(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        include_formulas: bool = False,
        include_formatting: bool = False
    ) -> str:
        """
        Async version of _run for better performance in async contexts.
        
        This method provides an asynchronous interface for the Excel reading operation,
        allowing it to be used efficiently in async agent execution flows. Currently
        delegates to the synchronous _run method, but structured for future async optimization.
        
        Args:
            file_path: Path to Excel file or base64-encoded file data
            sheet_name: Optional sheet name to read
            include_formulas: If True, includes cell formulas in output
            include_formatting: If True, includes formatting information
        
        Returns:
            str: JSON string containing structured Excel data, same as _run().
        
        Note:
            Currently implemented as a wrapper around _run() for compatibility.
            Future versions may implement true async file I/O for better performance.
        """
        return self._run(file_path, sheet_name, include_formulas, include_formatting)


class ModifyExcelInput(BaseModel):
    """Input schema for ModifyExcelTool."""
    file_path: str = Field(..., description="Path to Excel file or base64-encoded file data")
    operations: List[Dict[str, Any]] = Field(..., description="List of operations to perform on the Excel file")
    output_format: str = Field("base64", description="Output format: 'base64' or 'file_path' (default: 'base64')")


class ModifyExcelTool(BaseTool):
    """
    Modify Excel files by performing operations like setting cell values, adding formulas, and formatting cells.
    
    This tool enables comprehensive Excel file modification through a list of operations. It supports
    setting individual cells or ranges, adding formulas, applying formatting (fonts, colors, alignment),
    and modifying spreadsheet structure (adding rows/columns). The tool preserves the original file
    structure while applying requested changes.
    
    **When to use this tool:**
    - Updating Excel file data programmatically
    - Adding formulas to cells (e.g., SUM, AVERAGE, IF statements)
    - Formatting cells (bold, italic, colors, alignment)
    - Modifying Excel structure (adding rows or columns)
    - Creating or modifying Excel reports
    - Applying conditional formatting or styling
    - Building Excel files from structured data
    
    **When NOT to use this tool:**
    - For reading Excel files (use ReadExcelStructuredTool)
    - For creating charts (use CreateExcelChartTool)
    - For simple data export (consider CSV or JSON formats)
    - When you need to delete data (operations don't support deletion yet)
    
    **Input requirements:**
    - Must provide a valid Excel file (.xlsx format) as base64 or file path
    - Operations list must contain valid operation dictionaries
    - Each operation must have a "type" field specifying the operation type
    - Cell references must be valid Excel cell notation (e.g., "A1", "B2")
    
    **Output:**
    - Returns JSON with status, base64-encoded modified file, file name, and MIME type
    - Modified file is ready for download or further processing
    - Returns error message if modification fails
    
    **Supported Operations:**
    - set_cell: Set value in a specific cell (e.g., {"type": "set_cell", "cell": "A1", "value": "Hello"})
    - set_range: Set values in a range starting from a cell (2D array)
    - add_formula: Add formula to a cell (e.g., {"type": "add_formula", "cell": "C1", "formula": "=A1+B1"})
    - format_cell: Format cell appearance (font, fill, alignment)
    - add_row: Add a new row with data
    - add_column: Add a new column with data
    
    **Limitations:**
    - Operations are applied sequentially (order matters)
    - Cannot delete cells or rows (only add/modify)
    - Formatting options are limited to basic styles
    - Large files with many operations may take time to process
    
    **Example use cases:**
    1. "Add a formula to cell C10 that sums A1:A9"
    2. "Format all cells in column A as bold with yellow background"
    3. "Set the value in cell B5 to 'Total' and make it bold"
    4. "Add a new row at row 10 with data [1, 2, 3, 4, 5]"
    
    **Configuration:**
    Requires openpyxl library to be installed. The tool automatically handles both base64-encoded
    files and file paths. Output is always base64-encoded for easy integration with Streamlit downloads.
    """
    name: str = "modify_excel_file"
    description: str = (
        "Modify Excel files by performing operations like setting cell values, adding formulas, and formatting cells. "
        "Operations include: set_cell, set_range, add_formula, format_cell, add_row, add_column. "
        "Returns base64-encoded modified Excel file ready for download. "
        "IMPORTANT: File can be base64-encoded string or file path. Operations are applied sequentially. "
        "Best for: updating Excel data, adding formulas, applying formatting, modifying spreadsheet structure. "
        "NOT suitable for: reading files, creating charts, or deleting data."
    )
    args_schema: Type[BaseModel] = ModifyExcelInput

    def _run(
        self,
        file_path: str,
        operations: List[Dict[str, Any]],
        output_format: str = "base64"
    ) -> str:
        """
        Modify Excel file with operations.
        
        This method performs the Excel file modification:
        1. Decodes file input (base64 or file path)
        2. Loads workbook using openpyxl
        3. Applies each operation in sequence
        4. Saves modified workbook to bytes
        5. Returns base64-encoded file or saves to path
        
        Args:
            file_path: Path to Excel file or base64-encoded file data.
                      Example: "data.xlsx" or base64 string from file upload
            operations: List of operation dictionaries. Each operation must have:
                      - "type": Operation type (set_cell, set_range, add_formula, format_cell, add_row, add_column)
                      - Operation-specific fields (cell, value, formula, format, etc.)
                      Example: [{"type": "set_cell", "cell": "A1", "value": "Hello"}]
            output_format: Output format - "base64" (default) or "file_path"
        
        Returns:
            str: JSON string with status, base64-encoded file data, file name, and MIME type.
                 Returns error message string if modification fails.
        
        Raises:
            No exceptions are raised - all errors are caught and returned as error message strings.
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl library not available")
            return "Error: openpyxl library is not installed. Please install it with: pip install openpyxl"
        
        try:
            logger.info(f"Starting Excel file modification with {len(operations)} operations")
            logger.debug(f"Operations: {json.dumps(operations, indent=2) if operations else '[]'}")
            logger.debug(f"Output format: {output_format}")
            
            # Decode file input
            try:
                logger.debug("Decoding file input")
                file_bytes, is_base64 = _decode_file_input(file_path)
                logger.info(f"File decoded successfully (size: {len(file_bytes)} bytes)")
            except ValueError as decode_error:
                logger.error(f"File input decoding failed: {str(decode_error)}", exc_info=True)
                return f"Error: Invalid file input - {str(decode_error)}"
            except Exception as decode_error:
                logger.error(f"Unexpected error during file decoding: {str(decode_error)}", exc_info=True)
                return f"Error: Failed to decode file input - {str(decode_error)}"
            
            # Create BytesIO object
            try:
                logger.debug("Creating BytesIO object")
                file_obj = io.BytesIO(file_bytes)
                logger.debug("BytesIO object created")
            except Exception as io_error:
                logger.error(f"Error creating BytesIO object: {str(io_error)}", exc_info=True)
                return f"Error: Failed to create file object - {str(io_error)}"
            
            # Load workbook
            try:
                logger.debug("Loading workbook")
                wb = load_workbook(file_obj)
                logger.info(f"Workbook loaded successfully (sheets: {len(wb.sheetnames)})")
            except Exception as load_error:
                logger.error(f"Error loading workbook: {str(load_error)}", exc_info=True)
                return f"Error: Failed to load Excel workbook - {str(load_error)}"
            
            # Select active sheet
            try:
                ws = wb.active  # Use active sheet, can be extended to support sheet selection
                logger.debug(f"Using active sheet: {ws.title}")
            except Exception as sheet_error:
                logger.error(f"Error accessing active sheet: {str(sheet_error)}", exc_info=True)
                return f"Error: Failed to access worksheet - {str(sheet_error)}"
            
            # Perform operations
            operations_succeeded = 0
            operations_failed = 0
            
            for op_idx, op in enumerate(operations, 1):
                try:
                    logger.debug(f"Processing operation {op_idx}/{len(operations)}: {op.get('type', 'unknown')}")
                    op_type = op.get("type")
                    
                    if not op_type:
                        logger.warning(f"Operation {op_idx} missing 'type' field, skipping")
                        operations_failed += 1
                        continue
                    
                    if op_type == "set_cell":
                        try:
                            sheet = wb[op.get("sheet", ws.title)]
                            cell_ref = op.get("cell")
                            value = op.get("value")
                            logger.debug(f"Setting cell {cell_ref} = {value}")
                            sheet[cell_ref] = value
                            operations_succeeded += 1
                            logger.debug(f"Successfully set cell {cell_ref}")
                        except Exception as set_cell_error:
                            logger.error(f"Error in set_cell operation {op_idx}: {str(set_cell_error)}", exc_info=True)
                            operations_failed += 1
                            continue
                    
                    elif op_type == "set_range":
                        try:
                            sheet = wb[op.get("sheet", ws.title)]
                            start_cell = op.get("start_cell")
                            values = op.get("values")  # 2D list
                            logger.debug(f"Setting range starting at {start_cell} with {len(values)} rows")
                            for i, row in enumerate(values):
                                for j, val in enumerate(row):
                                    cell = sheet.cell(
                                        row=sheet[start_cell].row + i,
                                        column=sheet[start_cell].column + j
                                    )
                                    cell.value = val
                            operations_succeeded += 1
                            logger.debug(f"Successfully set range starting at {start_cell}")
                        except Exception as set_range_error:
                            logger.error(f"Error in set_range operation {op_idx}: {str(set_range_error)}", exc_info=True)
                            operations_failed += 1
                            continue
                    
                    elif op_type == "add_formula":
                        try:
                            sheet = wb[op.get("sheet", ws.title)]
                            cell_ref = op.get("cell")
                            formula = op.get("formula")
                            logger.debug(f"Adding formula to {cell_ref}: {formula}")
                            sheet[cell_ref] = formula
                            operations_succeeded += 1
                            logger.debug(f"Successfully added formula to {cell_ref}")
                        except Exception as formula_error:
                            logger.error(f"Error in add_formula operation {op_idx}: {str(formula_error)}", exc_info=True)
                            operations_failed += 1
                            continue
                    
                    elif op_type == "format_cell":
                        try:
                            sheet = wb[op.get("sheet", ws.title)]
                            cell_ref = op.get("cell")
                            format_config = op.get("format", {})
                            cell = sheet[cell_ref]
                            logger.debug(f"Formatting cell {cell_ref} with config: {format_config}")
                            
                            # Font formatting
                            if "font" in format_config:
                                try:
                                    font_config = format_config["font"]
                                    from openpyxl.styles import Font
                                    cell.font = Font(
                                        bold=font_config.get("bold", False),
                                        italic=font_config.get("italic", False),
                                        size=font_config.get("size"),
                                        name=font_config.get("name")
                                    )
                                    logger.debug(f"Applied font formatting to {cell_ref}")
                                except Exception as font_error:
                                    logger.warning(f"Error applying font formatting to {cell_ref}: {str(font_error)}")
                            
                            # Fill formatting
                            if "fill" in format_config:
                                try:
                                    fill_config = format_config["fill"]
                                    from openpyxl.styles import PatternFill
                                    if "fgColor" in fill_config:
                                        cell.fill = PatternFill(start_color=fill_config["fgColor"], 
                                                               end_color=fill_config["fgColor"], 
                                                               fill_type="solid")
                                        logger.debug(f"Applied fill formatting to {cell_ref}")
                                except Exception as fill_error:
                                    logger.warning(f"Error applying fill formatting to {cell_ref}: {str(fill_error)}")
                            
                            # Alignment
                            if "alignment" in format_config:
                                try:
                                    align_config = format_config["alignment"]
                                    from openpyxl.styles import Alignment
                                    cell.alignment = Alignment(
                                        horizontal=align_config.get("horizontal", "general"),
                                        vertical=align_config.get("vertical", "bottom")
                                    )
                                    logger.debug(f"Applied alignment formatting to {cell_ref}")
                                except Exception as align_error:
                                    logger.warning(f"Error applying alignment formatting to {cell_ref}: {str(align_error)}")
                            
                            operations_succeeded += 1
                            logger.debug(f"Successfully formatted cell {cell_ref}")
                        except Exception as format_error:
                            logger.error(f"Error in format_cell operation {op_idx}: {str(format_error)}", exc_info=True)
                            operations_failed += 1
                            continue
                    
                    elif op_type == "add_row":
                        try:
                            sheet = wb[op.get("sheet", ws.title)]
                            row_data = op.get("row_data", [])
                            row_num = op.get("row", sheet.max_row + 1)
                            logger.debug(f"Adding row {row_num} with {len(row_data)} columns")
                            for col_idx, value in enumerate(row_data, start=1):
                                sheet.cell(row=row_num, column=col_idx, value=value)
                            operations_succeeded += 1
                            logger.debug(f"Successfully added row {row_num}")
                        except Exception as add_row_error:
                            logger.error(f"Error in add_row operation {op_idx}: {str(add_row_error)}", exc_info=True)
                            operations_failed += 1
                            continue
                    
                    elif op_type == "add_column":
                        try:
                            sheet = wb[op.get("sheet", ws.title)]
                            col_data = op.get("column_data", [])
                            col_letter = op.get("column", chr(65 + sheet.max_column))  # Default to next column
                            logger.debug(f"Adding column {col_letter} with {len(col_data)} rows")
                            for row_idx, value in enumerate(col_data, start=1):
                                sheet[f"{col_letter}{row_idx}"] = value
                            operations_succeeded += 1
                            logger.debug(f"Successfully added column {col_letter}")
                        except Exception as add_col_error:
                            logger.error(f"Error in add_column operation {op_idx}: {str(add_col_error)}", exc_info=True)
                            operations_failed += 1
                            continue
                    else:
                        logger.warning(f"Unknown operation type '{op_type}' in operation {op_idx}, skipping")
                        operations_failed += 1
                        continue
                        
                except Exception as op_error:
                    logger.error(f"Unexpected error processing operation {op_idx}: {str(op_error)}", exc_info=True)
                    operations_failed += 1
                    continue
            
            logger.info(f"Operations completed: {operations_succeeded} succeeded, {operations_failed} failed")
            
            # Save to bytes
            try:
                logger.debug("Saving workbook to bytes")
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                file_bytes = output.read()
                logger.info(f"Workbook saved successfully (size: {len(file_bytes)} bytes)")
            except Exception as save_error:
                logger.error(f"Error saving workbook: {str(save_error)}", exc_info=True)
                return f"Error: Failed to save workbook - {str(save_error)}"
            
            # Encode or save to file
            try:
                if output_format == "base64":
                    logger.debug("Encoding file to base64")
                    encoded = _encode_file_output(file_bytes)
                    logger.info("Successfully modified Excel file")
                    return json.dumps({
                        "status": "success",
                        "file_data": encoded,
                        "file_name": "modified.xlsx",
                        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "operations_succeeded": operations_succeeded,
                        "operations_failed": operations_failed
                    })
                else:
                    logger.debug(f"Saving file to path: {output_format}")
                    with open(output_format, 'wb') as f:
                        f.write(file_bytes)
                    logger.info(f"File saved to: {output_format}")
                    return json.dumps({
                        "status": "success",
                        "file_path": output_format,
                        "operations_succeeded": operations_succeeded,
                        "operations_failed": operations_failed
                    })
            except Exception as output_error:
                logger.error(f"Error in output formatting: {str(output_error)}", exc_info=True)
                return f"Error: Failed to format output - {str(output_error)}"
            
        except ValueError as val_error:
            logger.error(f"Value error in Excel modification: {str(val_error)}", exc_info=True)
            return f"Error: Invalid input - {str(val_error)}"
        except MemoryError as mem_error:
            logger.error(f"Memory error modifying Excel file (file may be too large): {str(mem_error)}", exc_info=True)
            return f"Error: File too large to process - {str(mem_error)}"
        except Exception as e:
            logger.error(f"Unexpected error modifying Excel file: {str(e)}", exc_info=True)
            return f"Error modifying Excel file: {str(e)}"
    
    async def _arun(
        self,
        file_path: str,
        operations: List[Dict[str, Any]],
        output_format: str = "base64"
    ) -> str:
        """
        Async version of _run for better performance in async contexts.
        
        This method provides an asynchronous interface for Excel file modification,
        allowing it to be used efficiently in async agent execution flows. Currently
        delegates to the synchronous _run method, but structured for future async optimization.
        
        Args:
            file_path: Path to Excel file or base64-encoded file data
            operations: List of operation dictionaries
            output_format: Output format - "base64" or "file_path"
        
        Returns:
            str: JSON string with modified file data, same as _run().
        
        Note:
            Currently implemented as a wrapper around _run() for compatibility.
            Future versions may implement true async file I/O for better performance.
        """
        return self._run(file_path, operations, output_format)


class CreateExcelChartInput(BaseModel):
    """Input schema for CreateExcelChartTool."""
    file_path: str = Field(..., description="Path to Excel file or base64-encoded file data")
    chart_config: Dict[str, Any] = Field(..., description="Chart configuration dictionary")
    output_format: str = Field("base64", description="Output format: 'base64' or 'file_path' (default: 'base64')")


class CreateExcelChartTool(BaseTool):
    """
    Create charts in Excel files (bar, line, pie, scatter, etc.).
    
    This tool creates visualizations directly embedded in Excel files. It reads data from a specified
    range in the Excel file and creates a chart of the requested type, positioning it on the same or
    different sheet. Charts are fully integrated with Excel and can be edited using Excel's native tools.
    
    **When to use this tool:**
    - Adding visualizations to Excel reports
    - Creating charts from existing Excel data
    - Enhancing Excel files with data visualizations
    - Creating dashboard-style Excel files with charts
    - Generating Excel reports with embedded charts
    - Visualizing trends and comparisons in Excel data
    
    **When NOT to use this tool:**
    - For creating standalone charts (use visualization tools instead)
    - When you need interactive charts (Excel charts are static)
    - For complex chart customizations (use Excel directly)
    - When data is not in Excel format (convert first)
    
    **Input requirements:**
    - Must provide a valid Excel file (.xlsx format) as base64 or file path
    - Chart configuration must include:
      - sheet: Sheet name containing data
      - chart_type: Type of chart (bar, line, pie, scatter)
      - data_range: Data range in Excel notation (e.g., "A1:B10")
      - title: Chart title
      - position: Chart position on sheet (e.g., "E2")
    - Optional: x_axis_title, y_axis_title for axis labels
    
    **Output:**
    - Returns JSON with status, base64-encoded Excel file with chart, file name, and MIME type
    - Chart is embedded in the Excel file and can be edited in Excel
    - Returns error message if chart creation fails
    
    **Supported Chart Types:**
    - bar: Bar chart (horizontal bars)
    - line: Line chart (trend visualization)
    - pie: Pie chart (part-to-whole relationships)
    - scatter: Scatter plot (correlation analysis)
    
    **Limitations:**
    - Chart types are limited to basic types (bar, line, pie, scatter)
    - Complex chart customizations not supported
    - Data range must be valid and contain numeric data
    - Charts are static (not interactive)
    
    **Example use cases:**
    1. "Create a bar chart from data in range A1:B10 and place it at E2"
    2. "Add a line chart showing sales trends to the Excel file"
    3. "Create a pie chart from the data in the 'Summary' sheet"
    4. "Generate a scatter plot comparing two columns in the Excel file"
    
    **Configuration:**
    Requires openpyxl library to be installed. The tool automatically handles both base64-encoded
    files and file paths. Charts are embedded using openpyxl's chart API.
    """
    name: str = "create_excel_chart"
    description: str = (
        "Create charts in Excel files (bar, line, pie, scatter, etc.). "
        "Returns base64-encoded Excel file with chart embedded. "
        "Supports various chart types and customization options. "
        "IMPORTANT: File can be base64-encoded string or file path. Data range must be valid Excel notation. "
        "Best for: adding visualizations to Excel reports, creating charts from Excel data, enhancing reports. "
        "NOT suitable for: standalone charts, interactive visualizations, or complex chart customizations."
    )
    args_schema: Type[BaseModel] = CreateExcelChartInput

    def _run(
        self,
        file_path: str,
        chart_config: Dict[str, Any],
        output_format: str = "base64"
    ) -> str:
        """
        Create chart in Excel file.
        
        This method performs the chart creation operation:
        1. Decodes file input (base64 or file path)
        2. Loads workbook using openpyxl
        3. Parses data range and extracts data
        4. Creates chart object based on chart type
        5. Configures chart (title, axes, labels)
        6. Adds chart to sheet at specified position
        7. Saves workbook and returns base64-encoded file
        
        Args:
            file_path: Path to Excel file or base64-encoded file data.
                      Example: "data.xlsx" or base64 string from file upload
            chart_config: Dictionary containing chart configuration:
                        - sheet: Sheet name (default: active sheet)
                        - chart_type: Type of chart (bar, line, pie, scatter)
                        - data_range: Data range (e.g., "A1:B10")
                        - title: Chart title
                        - position: Chart position (e.g., "E2")
                        - x_axis_title: Optional X-axis title
                        - y_axis_title: Optional Y-axis title
            output_format: Output format - "base64" (default) or "file_path"
        
        Returns:
            str: JSON string with status, base64-encoded Excel file with chart, file name, and MIME type.
                 Returns error message string if chart creation fails.
        
        Raises:
            No exceptions are raised - all errors are caught and returned as error message strings.
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl library not available")
            return "Error: openpyxl library is not installed. Please install it with: pip install openpyxl"
        
        try:
            logger.info("Starting Excel chart creation")
            logger.debug(f"Chart config: {json.dumps(chart_config, indent=2) if chart_config else '{}'}")
            logger.debug(f"Output format: {output_format}")
            
            # Decode file input
            try:
                logger.debug("Decoding file input")
                file_bytes, is_base64 = _decode_file_input(file_path)
                logger.info(f"File decoded successfully (size: {len(file_bytes)} bytes)")
            except ValueError as decode_error:
                logger.error(f"File input decoding failed: {str(decode_error)}", exc_info=True)
                return f"Error: Invalid file input - {str(decode_error)}"
            except Exception as decode_error:
                logger.error(f"Unexpected error during file decoding: {str(decode_error)}", exc_info=True)
                return f"Error: Failed to decode file input - {str(decode_error)}"
            
            # Create BytesIO object
            try:
                logger.debug("Creating BytesIO object")
                file_obj = io.BytesIO(file_bytes)
                logger.debug("BytesIO object created")
            except Exception as io_error:
                logger.error(f"Error creating BytesIO object: {str(io_error)}", exc_info=True)
                return f"Error: Failed to create file object - {str(io_error)}"
            
            # Load workbook
            try:
                logger.debug("Loading workbook")
                wb = load_workbook(file_obj)
                logger.info(f"Workbook loaded successfully (sheets: {len(wb.sheetnames)})")
            except Exception as load_error:
                logger.error(f"Error loading workbook: {str(load_error)}", exc_info=True)
                return f"Error: Failed to load Excel workbook - {str(load_error)}"
            
            # Select sheet
            try:
                sheet_name = chart_config.get("sheet", wb.active.title)
                logger.debug(f"Selecting sheet: {sheet_name}")
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
                    return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
                ws = wb[sheet_name]
                logger.debug(f"Sheet '{sheet_name}' selected successfully")
            except Exception as sheet_error:
                logger.error(f"Error selecting sheet: {str(sheet_error)}", exc_info=True)
                return f"Error: Failed to select sheet - {str(sheet_error)}"
            
            # Parse data range
            try:
                data_range = chart_config.get("data_range", "A1:B10")
                logger.debug(f"Parsing data range: {data_range}")
                
                if ":" not in data_range:
                    logger.error(f"Invalid data_range format: {data_range}")
                    return "Error: Invalid data_range format. Use format like 'A1:B10'"
                
                start_cell, end_cell = data_range.split(":")
                logger.debug(f"Data range: {start_cell} to {end_cell}")
                
                try:
                    data = Reference(ws, min_col=ws[start_cell].column, min_row=ws[start_cell].row,
                                   max_col=ws[end_cell].column, max_row=ws[end_cell].row)
                    categories = Reference(ws, min_col=ws[start_cell].column, min_row=ws[start_cell].row,
                                          max_col=ws[start_cell].column, max_row=ws[end_cell].row)
                    values = Reference(ws, min_col=ws[start_cell].column + 1, min_row=ws[start_cell].row,
                                      max_col=ws[end_cell].column, max_row=ws[end_cell].row)
                    logger.debug(f"Data references created successfully")
                except Exception as ref_error:
                    logger.error(f"Error creating data references: {str(ref_error)}", exc_info=True)
                    return f"Error: Failed to create data references - {str(ref_error)}. Check that cell range is valid."
            except Exception as range_error:
                logger.error(f"Error parsing data range: {str(range_error)}", exc_info=True)
                return f"Error: Failed to parse data range - {str(range_error)}"
            
            # Create chart based on type
            try:
                chart_type = chart_config.get("chart_type", "bar").lower()
                logger.debug(f"Creating chart type: {chart_type}")
                
                if chart_type == "bar":
                    chart = BarChart()
                elif chart_type == "line":
                    chart = LineChart()
                elif chart_type == "pie":
                    chart = PieChart()
                elif chart_type == "scatter":
                    chart = ScatterChart()
                else:
                    logger.warning(f"Unknown chart type '{chart_type}', using default bar chart")
                    chart = BarChart()  # Default
                
                logger.debug(f"Chart object created: {type(chart).__name__}")
            except Exception as chart_create_error:
                logger.error(f"Error creating chart object: {str(chart_create_error)}", exc_info=True)
                return f"Error: Failed to create chart - {str(chart_create_error)}"
            
            # Configure chart
            try:
                chart_title = chart_config.get("title", "Chart")
                logger.debug(f"Configuring chart: title='{chart_title}'")
                chart.title = chart_title
                
                try:
                    chart.add_data(values, titles_from_data=True)
                    chart.set_categories(categories)
                    logger.debug("Chart data and categories added")
                except Exception as data_error:
                    logger.error(f"Error adding data to chart: {str(data_error)}", exc_info=True)
                    return f"Error: Failed to add data to chart - {str(data_error)}"
                
                if chart_config.get("x_axis_title"):
                    try:
                        chart.x_axis.title = chart_config["x_axis_title"]
                        logger.debug(f"X-axis title set: {chart_config['x_axis_title']}")
                    except Exception as x_axis_error:
                        logger.warning(f"Error setting x-axis title: {str(x_axis_error)}")
                
                if chart_config.get("y_axis_title"):
                    try:
                        chart.y_axis.title = chart_config["y_axis_title"]
                        logger.debug(f"Y-axis title set: {chart_config['y_axis_title']}")
                    except Exception as y_axis_error:
                        logger.warning(f"Error setting y-axis title: {str(y_axis_error)}")
                
                logger.debug("Chart configuration completed")
            except Exception as config_error:
                logger.error(f"Error configuring chart: {str(config_error)}", exc_info=True)
                return f"Error: Failed to configure chart - {str(config_error)}"
            
            # Add chart to sheet
            try:
                position = chart_config.get("position", "E2")
                logger.debug(f"Adding chart to sheet at position: {position}")
                ws.add_chart(chart, position)
                logger.debug("Chart added to sheet successfully")
            except Exception as add_chart_error:
                logger.error(f"Error adding chart to sheet: {str(add_chart_error)}", exc_info=True)
                return f"Error: Failed to add chart to sheet - {str(add_chart_error)}"
            
            # Save to bytes
            try:
                logger.debug("Saving workbook to bytes")
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                file_bytes = output.read()
                logger.info(f"Workbook saved successfully (size: {len(file_bytes)} bytes)")
            except Exception as save_error:
                logger.error(f"Error saving workbook: {str(save_error)}", exc_info=True)
                return f"Error: Failed to save workbook - {str(save_error)}"
            
            # Encode or save to file
            try:
                if output_format == "base64":
                    logger.debug("Encoding file to base64")
                    encoded = _encode_file_output(file_bytes)
                    logger.info("Successfully created Excel chart")
                    return json.dumps({
                        "status": "success",
                        "file_data": encoded,
                        "file_name": "chart.xlsx",
                        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    })
                else:
                    logger.debug(f"Saving file to path: {output_format}")
                    try:
                        with open(output_format, 'wb') as f:
                            f.write(file_bytes)
                        logger.info(f"File saved to: {output_format}")
                        return json.dumps({"status": "success", "file_path": output_format})
                    except PermissionError as perm_error:
                        logger.error(f"Permission denied writing to file: {output_format}")
                        return f"Error: Permission denied writing to file - {str(perm_error)}"
                    except Exception as file_write_error:
                        logger.error(f"Error writing file: {str(file_write_error)}", exc_info=True)
                        return f"Error: Failed to write file - {str(file_write_error)}"
            except Exception as output_error:
                logger.error(f"Error in output formatting: {str(output_error)}", exc_info=True)
                return f"Error: Failed to format output - {str(output_error)}"
            
        except ValueError as val_error:
            logger.error(f"Value error in Excel chart creation: {str(val_error)}", exc_info=True)
            return f"Error: Invalid input - {str(val_error)}"
        except MemoryError as mem_error:
            logger.error(f"Memory error creating Excel chart (file may be too large): {str(mem_error)}", exc_info=True)
            return f"Error: File too large to process - {str(mem_error)}"
        except Exception as e:
            logger.error(f"Unexpected error creating Excel chart: {str(e)}", exc_info=True)
            return f"Error creating Excel chart: {str(e)}"
    
    async def _arun(
        self,
        file_path: str,
        chart_config: Dict[str, Any],
        output_format: str = "base64"
    ) -> str:
        """
        Async version of _run for better performance in async contexts.
        
        This method provides an asynchronous interface for chart creation,
        allowing it to be used efficiently in async agent execution flows. Currently
        delegates to the synchronous _run method, but structured for future async optimization.
        
        Args:
            file_path: Path to Excel file or base64-encoded file data
            chart_config: Dictionary containing chart configuration
            output_format: Output format - "base64" or "file_path"
        
        Returns:
            str: JSON string with Excel file containing chart, same as _run().
        
        Note:
            Currently implemented as a wrapper around _run() for compatibility.
            Future versions may implement true async file I/O for better performance.
        """
        return self._run(file_path, chart_config, output_format)


class AnalyzeExcelSecurityInput(BaseModel):
    """Input schema for AnalyzeExcelSecurityTool."""
    file_path: str = Field(..., description="Path to Excel file or base64-encoded file data")
    checks: Optional[List[str]] = Field(None, description="Specific security checks to perform (optional)")


class AnalyzeExcelSecurityTool(BaseTool):
    """
    Analyze Excel files for security issues (macros, external links, hidden sheets, etc.).
    
    This tool performs comprehensive security analysis on Excel files, detecting potential threats
    and security risks. It checks for VBA macros, external links, hidden sheets, embedded objects,
    and suspicious formulas that could pose security risks. The analysis helps identify files that
    may contain malicious code or pose compliance risks.
    
    **When to use this tool:**
    - Performing security audits on Excel files
    - Detecting potential threats before opening files
    - Compliance checking for file security policies
    - Analyzing suspicious Excel files
    - Pre-screening files from untrusted sources
    - Security assessment of Excel-based reports
    - Identifying files with macros or external dependencies
    
    **When NOT to use this tool:**
    - For reading Excel file content (use ReadExcelStructuredTool)
    - For modifying Excel files (use ModifyExcelTool)
    - When you trust the file source completely
    - For performance analysis (this is security-focused)
    
    **Input requirements:**
    - Must provide a valid Excel file (.xlsx or .xls format) as base64 or file path
    - Optional: List of specific security checks to perform
    - If checks list is provided, only those checks are performed
    
    **Output:**
    - Returns JSON security analysis report with:
      - Security findings for each check performed
      - Risk level assessment (low, medium, high)
      - Recommendations for safe handling
      - Details of detected issues (macros, links, hidden sheets)
    - Returns error message if analysis fails
    
    **Security Checks Performed:**
    - macros: Detects VBA macros and embedded code
    - external_links: Identifies formulas linking to external files
    - hidden_sheets: Finds hidden worksheets
    - embedded_objects: Detects embedded objects (future enhancement)
    - password_protection: Checks for password protection (future enhancement)
    
    **Limitations:**
    - Cannot detect all types of malicious code (heuristic-based)
    - Some legitimate uses may trigger warnings (false positives)
    - Advanced obfuscation techniques may not be detected
    - Analysis is based on file structure, not code execution
    
    **Example use cases:**
    1. "Analyze this Excel file for security issues before I open it"
    2. "Check if this file contains macros or external links"
    3. "Perform a security audit on the uploaded Excel file"
    4. "Identify any hidden sheets or suspicious content in this spreadsheet"
    
    **Configuration:**
    Requires openpyxl library to be installed. The tool loads workbooks with VBA preservation
    enabled to detect macros. Analysis is read-only and does not modify the file.
    """
    name: str = "analyze_excel_security"
    description: str = (
        "Analyze Excel files for security issues including macros, external links, hidden sheets, "
        "embedded objects, and suspicious formulas. Returns security analysis report with findings and recommendations. "
        "IMPORTANT: File can be base64-encoded string or file path. Use this before opening files from untrusted sources. "
        "Best for: security audits, threat detection, compliance checking, pre-screening suspicious files. "
        "NOT suitable for: reading file content, modifying files, or performance analysis."
    )
    args_schema: Type[BaseModel] = AnalyzeExcelSecurityInput

    def _run(
        self,
        file_path: str,
        checks: Optional[List[str]] = None
    ) -> str:
        """
        Analyze Excel file for security issues.
        
        This method performs the security analysis:
        1. Decodes file input (base64 or file path)
        2. Loads workbook with VBA preservation enabled
        3. Performs security checks (macros, links, hidden sheets)
        4. Assesses overall risk level
        5. Generates recommendations
        6. Returns structured security report
        
        Args:
            file_path: Path to Excel file or base64-encoded file data.
                      Example: "data.xlsx" or base64 string from file upload
            checks: Optional list of specific checks to perform. If not provided, all checks are performed.
                   Valid check names: "macros", "external_links", "hidden_sheets"
                   Example: ["macros", "external_links"]
        
        Returns:
            str: JSON string containing security analysis report with:
                 - Security findings for each check
                 - Risk level (low, medium, high)
                 - Recommendations
                 - Details of detected issues
                 Returns error message string if analysis fails.
        
        Raises:
            No exceptions are raised - all errors are caught and returned as error message strings.
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl library not available")
            return "Error: openpyxl library is not installed. Please install it with: pip install openpyxl"
        
        try:
            logger.info("Starting Excel security analysis")
            logger.debug(f"Security checks requested: {checks if checks else 'all'}")
            logger.debug(f"File input length: {len(file_path) if file_path else 0}")
            
            # Decode file input
            try:
                logger.debug("Decoding file input")
                file_bytes, is_base64 = _decode_file_input(file_path)
                logger.info(f"File decoded successfully (size: {len(file_bytes)} bytes)")
            except ValueError as decode_error:
                logger.error(f"File input decoding failed: {str(decode_error)}", exc_info=True)
                return f"Error: Invalid file input - {str(decode_error)}"
            except Exception as decode_error:
                logger.error(f"Unexpected error during file decoding: {str(decode_error)}", exc_info=True)
                return f"Error: Failed to decode file input - {str(decode_error)}"
            
            # Create BytesIO object
            try:
                logger.debug("Creating BytesIO object")
                file_obj = io.BytesIO(file_bytes)
                logger.debug("BytesIO object created")
            except Exception as io_error:
                logger.error(f"Error creating BytesIO object: {str(io_error)}", exc_info=True)
                return f"Error: Failed to create file object - {str(io_error)}"
            
            # Load workbook with VBA preservation
            try:
                logger.debug("Loading workbook with VBA preservation (keep_vba=True)")
                wb = load_workbook(file_obj, keep_vba=True)  # Keep VBA for macro detection
                logger.info(f"Workbook loaded successfully (sheets: {len(wb.sheetnames)})")
                logger.debug(f"Available sheets: {', '.join(wb.sheetnames)}")
            except Exception as load_error:
                logger.error(f"Error loading workbook: {str(load_error)}", exc_info=True)
                return f"Error: Failed to load Excel workbook - {str(load_error)}"
            
            # Initialize report
            try:
                report = {
                    "file_analyzed": "Excel file",
                    "security_findings": [],
                    "risk_level": "low",
                    "recommendations": []
                }
                logger.debug("Security report initialized")
            except Exception as init_error:
                logger.error(f"Error initializing report: {str(init_error)}", exc_info=True)
                return f"Error: Failed to initialize security report - {str(init_error)}"
            
            # Determine which checks to perform
            checks_to_perform = checks if checks else ["macros", "external_links", "hidden_sheets"]
            logger.debug(f"Performing security checks: {checks_to_perform}")
            
            # Check for macros
            if "macros" in checks_to_perform:
                try:
                    logger.debug("Checking for VBA macros")
                    if hasattr(wb, 'vba_archive') and wb.vba_archive:
                        logger.warning("VBA macros detected in file")
                        report["security_findings"].append({
                            "check": "macros",
                            "status": "warning",
                            "message": "File contains VBA macros",
                            "risk": "medium"
                        })
                        report["risk_level"] = "medium"
                        logger.info("Macro check: WARNING - VBA macros found")
                    else:
                        logger.debug("No VBA macros detected")
                        report["security_findings"].append({
                            "check": "macros",
                            "status": "safe",
                            "message": "No VBA macros detected"
                        })
                        logger.info("Macro check: SAFE - No macros found")
                except AttributeError as attr_error:
                    logger.warning(f"Error checking for macros (attribute error): {str(attr_error)}")
                    report["security_findings"].append({
                        "check": "macros",
                        "status": "error",
                        "message": f"Could not check for macros: {str(attr_error)}"
                    })
                except Exception as macro_error:
                    logger.error(f"Error checking for macros: {str(macro_error)}", exc_info=True)
                    report["security_findings"].append({
                        "check": "macros",
                        "status": "error",
                        "message": f"Macro check failed: {str(macro_error)}"
                    })
            
            # Check for hidden sheets
            if "hidden_sheets" in checks_to_perform:
                try:
                    logger.debug("Checking for hidden sheets")
                    hidden_sheets = []
                    for sheet_name in wb.sheetnames:
                        try:
                            if wb[sheet_name].sheet_state == 'hidden':
                                hidden_sheets.append(sheet_name)
                                logger.debug(f"Found hidden sheet: {sheet_name}")
                        except Exception as sheet_check_error:
                            logger.warning(f"Error checking sheet state for '{sheet_name}': {str(sheet_check_error)}")
                            continue
                    
                    if hidden_sheets:
                        logger.warning(f"Found {len(hidden_sheets)} hidden sheet(s)")
                        report["security_findings"].append({
                            "check": "hidden_sheets",
                            "status": "warning",
                            "message": f"Found {len(hidden_sheets)} hidden sheet(s): {', '.join(hidden_sheets)}",
                            "risk": "low"
                        })
                        logger.info(f"Hidden sheets check: WARNING - {len(hidden_sheets)} hidden sheets found")
                    else:
                        logger.debug("No hidden sheets found")
                        report["security_findings"].append({
                            "check": "hidden_sheets",
                            "status": "safe",
                            "message": "No hidden sheets found"
                        })
                        logger.info("Hidden sheets check: SAFE - No hidden sheets")
                except Exception as hidden_sheets_error:
                    logger.error(f"Error checking for hidden sheets: {str(hidden_sheets_error)}", exc_info=True)
                    report["security_findings"].append({
                        "check": "hidden_sheets",
                        "status": "error",
                        "message": f"Hidden sheets check failed: {str(hidden_sheets_error)}"
                    })
            
            # Check for external links
            if "external_links" in checks_to_perform:
                try:
                    logger.debug("Checking for external links in formulas")
                    external_links = []
                    sheets_checked = 0
                    cells_checked = 0
                    
                    for sheet in wb.worksheets:
                        try:
                            sheets_checked += 1
                            logger.debug(f"Checking sheet '{sheet.title}' for external links")
                            for row in sheet.iter_rows():
                                for cell in row:
                                    cells_checked += 1
                                    try:
                                        if cell.data_type == 'f' and cell.value:
                                            formula = str(cell.value)
                                            # Check for external references (format: [filename]sheet!cell)
                                            if '[' in formula and ']' in formula:
                                                external_links.append({
                                                    "sheet": sheet.title,
                                                    "cell": cell.coordinate,
                                                    "formula": formula
                                                })
                                                logger.debug(f"External link found in {sheet.title}!{cell.coordinate}")
                                    except Exception as cell_error:
                                        logger.debug(f"Error checking cell {cell.coordinate if 'cell' in locals() else 'unknown'}: {str(cell_error)}")
                                        continue
                        except Exception as sheet_error:
                            logger.warning(f"Error checking sheet '{sheet.title}': {str(sheet_error)}")
                            continue
                    
                    logger.debug(f"External links check complete: {sheets_checked} sheets, {cells_checked} cells checked")
                    
                    if external_links:
                        logger.warning(f"Found {len(external_links)} external link(s)")
                        report["security_findings"].append({
                            "check": "external_links",
                            "status": "warning",
                            "message": f"Found {len(external_links)} external link(s)",
                            "risk": "medium",
                            "details": external_links[:10]  # Limit details
                        })
                        if report["risk_level"] == "low":
                            report["risk_level"] = "medium"
                        logger.info(f"External links check: WARNING - {len(external_links)} external links found")
                    else:
                        logger.debug("No external links found")
                        report["security_findings"].append({
                            "check": "external_links",
                            "status": "safe",
                            "message": "No external links found"
                        })
                        logger.info("External links check: SAFE - No external links")
                except Exception as external_links_error:
                    logger.error(f"Error checking for external links: {str(external_links_error)}", exc_info=True)
                    report["security_findings"].append({
                        "check": "external_links",
                        "status": "error",
                        "message": f"External links check failed: {str(external_links_error)}"
                    })
            
            # Generate recommendations
            try:
                logger.debug("Generating security recommendations")
                if report["risk_level"] != "low":
                    report["recommendations"].append("Review macros and external links before opening")
                    report["recommendations"].append("Scan file with antivirus software")
                    logger.debug(f"Added {len(report['recommendations'])} recommendations")
                else:
                    report["recommendations"].append("File appears safe, but always exercise caution with files from untrusted sources")
                    logger.debug("Added default recommendation for low-risk file")
            except Exception as rec_error:
                logger.warning(f"Error generating recommendations: {str(rec_error)}")
                report["recommendations"] = ["Unable to generate recommendations"]
            
            # Serialize report to JSON
            try:
                logger.debug("Serializing security report to JSON")
                json_report = json.dumps(report, indent=2)
                logger.info(f"Security analysis complete: {report['risk_level']} risk level, {len(report['security_findings'])} findings")
                logger.debug(f"JSON report length: {len(json_report)} characters")
                return json_report
            except Exception as json_error:
                logger.error(f"Error serializing report to JSON: {str(json_error)}", exc_info=True)
                return f"Error: Failed to serialize security report - {str(json_error)}"
            
        except ValueError as val_error:
            logger.error(f"Value error in Excel security analysis: {str(val_error)}", exc_info=True)
            return f"Error: Invalid input - {str(val_error)}"
        except MemoryError as mem_error:
            logger.error(f"Memory error analyzing Excel security (file may be too large): {str(mem_error)}", exc_info=True)
            return f"Error: File too large to process - {str(mem_error)}"
        except Exception as e:
            logger.error(f"Unexpected error analyzing Excel security: {str(e)}", exc_info=True)
            return f"Error analyzing Excel security: {str(e)}"
    
    async def _arun(
        self,
        file_path: str,
        checks: Optional[List[str]] = None
    ) -> str:
        """
        Async version of _run for better performance in async contexts.
        
        This method provides an asynchronous interface for security analysis,
        allowing it to be used efficiently in async agent execution flows. Currently
        delegates to the synchronous _run method, but structured for future async optimization.
        
        Args:
            file_path: Path to Excel file or base64-encoded file data
            checks: Optional list of specific security checks to perform
        
        Returns:
            str: JSON string containing security analysis report, same as _run().
        
        Note:
            Currently implemented as a wrapper around _run() for compatibility.
            Future versions may implement true async file I/O for better performance.
        """
        return self._run(file_path, checks)

